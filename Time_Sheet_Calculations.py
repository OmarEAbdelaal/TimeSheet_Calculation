import numpy as np
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from sqlalchemy import create_engine, text
from scipy.stats import ttest_ind
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook

def process_csv():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Ask user to select a CSV file
    file_path = filedialog.askopenfilename(title="Select CSV File", filetypes=[("CSV Files", "*.csv")])
    attendance_file = filedialog.askopenfilename(title="Select xlsx File", filetypes=[("xlsx Files", "*.xlsx")])

    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        return

    try:
        time_sheet = pd.read_csv(file_path)
        columns = ['User', 'Start Date', 'Start Time', 'End Time', 'Duration (h)', 'Task']
        time_data = time_sheet[columns]
        time_data['Start Date'] = pd.to_datetime(time_data['Start Date'])
        time_data['Duration (h)'] = pd.to_timedelta(time_data['Duration (h)'])
        time_data['Start Time'] = pd.to_datetime(time_data['Start Time']).dt.time
        time_data['End Time'] = pd.to_datetime(time_data['End Time']).dt.time

        # Filter and process 'Task' column
        time_data['Task'] = time_data['Task'].fillna('')
        filtered_data = time_data[time_data['Task'].str.contains('LEAVE')]
        time_data.loc[~time_data['Task'].str.contains('LEAVE'), 'Task'] = np.nan
        final_data = pd.concat([filtered_data, time_data[time_data['Task'].notna()]])
        final_data.reset_index(drop=True, inplace=True)

        grouped_time_data = time_data.groupby(['Start Date', 'User']).agg({
            'Start Time': 'min',
            'End Time': 'max',
            'Duration (h)': 'sum',
            'Task': 'sum'
        }).reset_index()

        unique_users = grouped_time_data['User'].unique()

        for user in unique_users:
            user_time_data = grouped_time_data[grouped_time_data['User'] == user]
            date_range = pd.date_range(start=user_time_data['Start Date'].min(), end=user_time_data['Start Date'].max(), freq='D')
            user_date_range_df = pd.DataFrame({'Start Date': date_range})

            merged_df = pd.merge(user_date_range_df, user_time_data, on='Start Date', how='left')
            merged_df['User'].fillna(user, inplace=True)
            merged_df.drop(columns=['User'], inplace=True)
            merged_df.replace(0, '', inplace=True)

            output_file = f'{user}.xlsx'
            merged_df.to_excel(output_file, index=False)

        
        destination_workbook = load_workbook(attendance_file)

        for user in unique_users:
            file_name = f"{user}.xlsx"
            if not os.path.exists(file_name):
                continue

            source_workbook = load_workbook(file_name)
            source_sheet = source_workbook.active
            destination_sheet = destination_workbook[user]

            source_range = source_sheet['A2:E32']
            for row in source_range:
                for cell in row:
                    destination_sheet[cell.coordinate].value = cell.value

        destination_workbook.save(attendance_file)
        messagebox.showinfo("Success", "Processing complete! Data saved to Excel.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

if __name__ == "__main__":
    process_csv()
