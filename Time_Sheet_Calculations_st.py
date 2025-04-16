import streamlit as st
import numpy as np
import pandas as pd
import os
from sqlalchemy import create_engine, text # Importing the SQL interface.
from scipy.stats import ttest_ind
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from io import BytesIO
import calendar

# Streamlit UI
st.title("Attendance Clockify Calculator")

# File uploaders
csv_file = st.file_uploader("Upload CSV File", type=["csv"])
attendance_file = "Attendance.xlsx"

if csv_file and attendance_file:
    try:
        # Load CSV data
        time_sheet = pd.read_csv(csv_file)
        
        # Process columns
        columns = ['User', 'Start Date', 'Start Time', 'End Time', 'Duration (h)', 'Task']
        time_data = time_sheet[columns]
        time_data['Start Date'] = pd.to_datetime(time_data['Start Date'])
        time_data['Duration (h)'] = pd.to_timedelta(time_data['Duration (h)'])
        time_data['Start Time'] = pd.to_datetime(time_data['Start Time']).dt.time
        time_data['End Time'] = pd.to_datetime(time_data['End Time']).dt.time

        # Process 'Task' column
        time_data['Task'] = time_data['Task'].fillna('')
        filtered_data = time_data[time_data['Task'].str.contains('LEAVE')]
        time_data.loc[~time_data['Task'].str.contains('LEAVE'), 'Task'] = np.nan
        final_data = pd.concat([filtered_data, time_data[time_data['Task'].notna()]]).reset_index(drop=True)

        # Group data
        grouped_time_data = time_data.groupby(['Start Date', 'User']).agg({
            'Start Time': 'min',
            'End Time': 'max',
            'Duration (h)': 'sum',
            'Task': 'sum'
        }).reset_index()

        unique_users = grouped_time_data['User'].unique()
        
        # Extract Year and Month from 'Start Date'
        Year = int(grouped_time_data['Start Date'].dt.year.unique()[0])
        Month = int(grouped_time_data['Start Date'].dt.month.unique()[0])
        Month_str = calendar.month_abbr[Month]
        
        # Calculating the working hours per month
        def calculate_working_days(year, month):
            # Generate all dates in the given month
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        
            # Create a date range
            date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        
            # Exclude Fridays (4) and Saturdays (5)
            working_days = [date for date in date_range if date.weekday() not in [4, 5]]
        
            return len(working_days)
        
        # Example Usage
        working_days_count = calculate_working_days(Year, Month)
        working_hours_per_month = working_days_count * 9
        time_format = timedelta(hours=working_hours_per_month)


        # Read the uploaded Excel file
        #attendance_bytes = BytesIO(attendance_file.read())
        destination_workbook = load_workbook()

        for user in unique_users:
            # Get the Last worksheet
            last_sheet = destination_workbook.worksheets[-1]

            # Rename the first worksheet
            last_sheet.title = f"{user}"

            # Duplicate the renamed sheet
            copied_sheet = destination_workbook.copy_worksheet(last_sheet)
            copied_sheet.title = "Template"
            
            user_time_data = grouped_time_data[grouped_time_data['User'] == user]
            date_range = pd.date_range(start=user_time_data['Start Date'].min(), end=user_time_data['Start Date'].max(), freq='D')
            user_date_range_df = pd.DataFrame({'Start Date': date_range})

            merged_df = pd.merge(user_date_range_df, user_time_data, on='Start Date', how='left')
            merged_df['User'].fillna(user, inplace=True)
            merged_df.drop(columns=['User'], inplace=True)
            merged_df.replace(0, '', inplace=True)

            # If sheet exists in destination workbook, update it
            if user in destination_workbook.sheetnames:
                destination_sheet = destination_workbook[user]

                for i, row in enumerate(merged_df.itertuples(index=False), start=2):
                    for j, value in enumerate(row, start=1):
                        destination_sheet.cell(row=i, column=j, value=value)
                        # Copy working hours into cell B33
                        destination_sheet["B33"].value = time_format

        # Save the updated Excel file
        output_stream = BytesIO()
        destination_workbook.save(output_stream)
        output_stream.seek(0)

        st.success("Processing complete! Click below to download the updated Excel file.")

        # Provide download button
        st.download_button(label="Download Updated Excel File",
                           data=output_stream,
                           file_name=f"{Month_str}._Attendance.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
